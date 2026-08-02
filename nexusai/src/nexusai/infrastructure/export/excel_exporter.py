"""Excel exporter, behind the optional ``excel`` extra.

Writes an ``.xlsx`` workbook with the dataset on one worksheet and its metadata,
validation and quality summaries on their own sheets, so a reviewer sees the data
and its quality side by side. openpyxl is imported lazily inside the method, so
importing this module never requires the optional dependency; only exporting does.

Excel's hard limits are respected rather than silently exceeded. If a dataset has
more rows or columns than a worksheet can hold, the export raises rather than
truncating -- silent truncation would present a partial dataset as complete. Every
cell passes through the formula-injection guard, since a spreadsheet is the whole
point of the format.
"""

from __future__ import annotations

from pathlib import Path

from nexusai.domain.errors.exceptions import ExportError
from nexusai.domain.model.persistence import DatasetVersion, ExportManifest
from nexusai.domain.model.processing import ProcessedDataset
from nexusai.infrastructure.export.rows import column_order, record_to_row
from nexusai.infrastructure.export.sanitize import neutralise
from nexusai.infrastructure.export.writer import (
    atomic_output,
    build_manifest,
    export_identity,
    resolve_target,
)

_VERSION = "1.0"
_MAX_ROWS = 1_048_576
_MAX_COLS = 16_384


class ExcelExporter:
    """Exports a dataset to a multi-sheet ``.xlsx`` workbook."""

    export_format = "xlsx"
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def __init__(self, base_dir: Path, *, include_source: bool = True) -> None:
        self._base = Path(base_dir)
        self._include_source = include_source

    def export(
        self,
        dataset: ProcessedDataset,
        destination: str,
        *,
        version: DatasetVersion | None = None,
    ) -> ExportManifest:
        """Write ``dataset`` as an Excel workbook and return the manifest.

        Raises:
            ExportError: If openpyxl is not installed, or the dataset exceeds
                Excel's row or column limits.
        """
        try:
            from openpyxl import Workbook
        except ImportError as exc:  # pragma: no cover - exercised via a stub
            raise ExportError("Excel export requires the 'excel' extra (openpyxl)") from exc

        columns = column_order(dataset, include_source=self._include_source)
        self._check_limits(dataset, columns)

        workbook = Workbook()
        data_sheet = workbook.active
        data_sheet.title = "Dataset"
        data_sheet.append(columns)
        for record in dataset.records:
            row = record_to_row(record, columns)
            data_sheet.append([neutralise(row[column]) for column in columns])

        self._write_summary_sheets(workbook, dataset)

        with atomic_output(self._base, destination) as temp:
            workbook.save(temp)
        dataset_id, dataset_version = export_identity(version)
        return build_manifest(
            resolve_target(self._base, destination),
            dataset_id=dataset_id,
            dataset_version=dataset_version,
            export_format=self.export_format,
            media_type=self.media_type,
            record_count=len(dataset.records),
            duration_seconds=0.0,
            exporter_version=_VERSION,
        )

    def _check_limits(self, dataset: ProcessedDataset, columns: list[str]) -> None:
        if len(dataset.records) + 1 > _MAX_ROWS:
            raise ExportError(
                "Dataset exceeds Excel's row limit; export would truncate",
                rows=len(dataset.records),
                limit=_MAX_ROWS,
            )
        if len(columns) > _MAX_COLS:
            raise ExportError(
                "Dataset exceeds Excel's column limit; export would truncate",
                columns=len(columns),
                limit=_MAX_COLS,
            )

    def _write_summary_sheets(self, workbook: object, dataset: ProcessedDataset) -> None:
        context = dataset.context
        if context is None:
            return
        metadata_sheet = workbook.create_sheet("Metadata")  # type: ignore[attr-defined]
        metadata_sheet.append(["key", "value"])
        metadata_sheet.append(["framework_version", context.framework_version])
        metadata_sheet.append(["rule_version", context.rule_version])
        metadata_sheet.append(["quality_grade", context.quality_grade.value])
        metadata_sheet.append(["record_count", len(dataset.records)])

        validation_sheet = workbook.create_sheet("Validation")  # type: ignore[attr-defined]
        validation_sheet.append(["code", "severity", "message", "location"])
        for issue in context.validation_summary.issues:
            validation_sheet.append(
                [
                    neutralise(issue.code),
                    issue.severity.name,
                    neutralise(issue.message),
                    issue.location or "",
                ]
            )

        quality_sheet = workbook.create_sheet("Quality")  # type: ignore[attr-defined]
        quality_sheet.append(["dimension", "score", "weight"])
        for measurement in context.quality.measurements:
            quality_sheet.append([measurement.dimension, measurement.score, measurement.weight])
